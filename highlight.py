import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# CSV 파일을 불러옵니다.
csv_file = 'concatenated.csv'
df = pd.read_csv(csv_file)

# 새로운 Excel 워크북을 생성합니다.
wb = Workbook()
ws = wb.active

# pandas DataFrame을 Excel 워크시트로 변환합니다.
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 하이라이트 처리할 셀의 색상을 설정합니다 (이 경우, 노란색).
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 각 열에 대해 평균값 이상인 값을 하이라이트 처리합니다.
for col in range(1, 103):  # 100개의 열에 대해 반복
    avg_cell = ws.cell(row=35, column=col)
    for row in range(2, 35):  # 제목 행 및 평균값 행 제외
        cell = ws.cell(row=row, column=col)
        if cell.value >= avg_cell.value:
            cell.fill = highlight_fill

# 결과를 새로운 Excel 파일로 저장합니다.
output_file = 'highlighted_data.xlsx'
wb.save(output_file)
